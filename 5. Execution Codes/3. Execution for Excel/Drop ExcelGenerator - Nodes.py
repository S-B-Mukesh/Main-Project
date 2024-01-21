from openpyxl import Workbook
import os
a = ["10-5"]

for file in a:
	for attack in ["constant", "periodic", "random"]:
		wb = Workbook()
		s = wb.active
		s.cell(1,1).value = "Malicious Nodes"
		s.cell(1,2).value = "Nodes"
		for i in range(10):
			s.cell(1,i+3).value = i

		percent=6
		p=2
		for f1 in range(percent):
			s.cell(p,1).value = f1
			for node in range(10):
				cd = [0]*10
				cp = open("G:/Results/10-5/Delay/"+attack[0]+"_Delay_0."+str(f1)+"0","r")
				for line in cp:
					parts = line.split()
					if parts[2] =="_"+str(node)+"_":
						if line[0]=="D":
							cd[int(parts[14][0])] += 1
				
				s.cell(p,2).value = node
				for i in range(10):
					s.cell(p,i+3).value = cd[i]
				p += 1
		wb.save("C:/Users/R&D LAB/Desktop/Temp/New Graphs/Dataset for packets drop for all Nodes in "+attack+" attack.xlsx")
		wb.close()