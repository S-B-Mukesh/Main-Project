import os
import matplotlib.pyplot as plt
import numpy as np
import openpyxl

for i in ["constant", "periodic", "random"]:
	barWidth = 0.15
	fig = plt.subplots(figsize =(15,10))
	w = openpyxl.load_workbook("I:/Today/Temp/New Graphs/PDeR "+i+".xlsx")
	s = w.active
	for Attacker in range(2,53,10):
		temp = (Attacker-2)//10
		x_axis = list(range(10))
		y_axis = []
		for j in range(Attacker,Attacker+10):
			print(s.cell(j,37).value,type(s.cell(j,37).value),j,37)
			if s.cell(j,37).value != "#DIV/0!":
				y_axis.append(s.cell(j,37).value)
			else:
				y_axis.append(0)
		f = plt.figure()
		f.set_figwidth(15)
		f.set_figheight(10)
		plt.plot(x_axis[:10-temp],y_axis[:10-temp], color='blue',label="Benign Nodes")
		plt.plot(x_axis[10-temp-1:],y_axis[10-temp-1:], color='red',label="Malicious NOdes")
		plt.xticks(x_axis)
		plt.yticks(y_axis)
		plt.title("Packet Drop ratio in "+i+" attack when "+str((Attacker-2)//10)+" Attackers", fontdict={'fontsize': 25})
		plt.xlabel("Nodes", fontdict={'fontsize': 25}) 
		plt.ylabel("Packet Drop Ratio", fontdict={'fontsize': 25})
		ax = plt.subplot()
		ax.tick_params(axis='x', labelsize=30)
		ax.tick_params(axis='y', labelsize=30)
		plt.legend()
		plt.savefig("I:/Today/Detection/10-5/"+i+"/D_S "+i+" "+str((Attacker-2)//10))
		plt.close()